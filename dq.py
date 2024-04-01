import pandas as pd 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font,Border,Side,PatternFill
from openpyxl.worksheet.views import SheetView
import os
import xlwings as xw
import pandas 
# import dask.dataframe as dd
from datetime import datetime
import time as tme
import shutil
import numpy as np
from openpyxl.chart import BarChart,Reference
#print(datetime.now())
def dq_validation(path):
    cur_date_time=datetime.now()
    cur_date_time_f=cur_date_time.strftime("%Y%m%d_%H%M%S")
    #print(len("DQ_Validation_Results"+time))
    path_1=path+'//'+'Input_Files'
    result_path=os.path.join(path,"Test Result")
    if not os.path.exists(result_path):
        os.makedirs(result_path)
    New_Val_Res_Path=os.path.join(result_path,"DQ_Validation_Results_"+cur_date_time_f)
    if not os.path.exists(New_Val_Res_Path):
        os.makedirs(New_Val_Res_Path)
    Log_Path=os.path.join(path,"DQ_Logs")
    if not os.path.exists(Log_Path):
        os.makedirs(Log_Path)
    encoding=['utf-8','latin-1','iso-8859-1','cp1252']
    
    file_log=[]
    file_error=[]
    for file_name in os.listdir(path_1):
        try:
            if file_name.endswith('.TXT') or file_name.endswith('.txt') or file_name.endswith('.csv'):
                dic_data_max={}
                dic_data_min={}
                tot_rec=0
                unique_status={}
                unique_count={}
                not_null_count={}
                alpha_numeric={}
                numeric_count={}
                delimiter='|'
                if file_name.endswith('.csv'):
                    delimiter=','
                main_df={'Scenarios':['Total No. Of Records','No. Of Unique Records','Max Length For Each Column','Min Length For Each Column','Not Null Count For Each Column','Null/Blank Count For Each Column','Percentage Of Data Population','Unique Column Check','AlphaNumeric Record Count','AlphaNumeric Check','Numeric Count']}
                print('yes')
                file_nm=file_name.split('.TXT')[0] if file_name.endswith('.TXT') else file_name.split('.csv')[0] if file_name.endswith('.csv') else file_name.split('.txt')[0] 
                if file_nm+'_config.xlsx' in os.listdir(path_1):
                    df_config=pd.read_excel(path_1+'//'+file_nm+'_config.xlsx')
                else:
                    df_config=pd.read_excel(path_1+'//'+file_nm[:-15]+'_config.xlsx')
                if df_config['Fixed_Width'].empty:
                    df=pd.read_csv(path_1+'//'+file_name,sep=delimiter,dtype=str,header=None,na_filter=False)
                    with open(path_1+'//'+file_name,'r', encoding='utf-8') as f:
                        lines=f.readlines()
                        last_line=lines[-1].strip()
                    if delimiter=='|':
                        if not last_line.endswith(delimiter):
                            df=df.iloc[:-1]      
                    #print('yes')
                    if df_config['Columns'].empty:
                        df.columns=df.iloc[0]
                        df=df.iloc[1:]
                    else:
                        df.columns=df_config['Columns'].dropna().tolist()
                    tot_rec=tot_rec+df.shape[0]
                    for i in df.columns: 
                        df[i]=df[i].replace('nan',np.nan)
                        max_length=df[i].dropna().astype(str).str.len().max()
                        if i not in  dic_data_max:
                            if str(max_length)=='nan':
                                dic_data_max[i]=0
                            else:
                                dic_data_max[i]=max_length
                        else:
                            if str(max_length)!='nan':
                                if dic_data_max[i]<max_length:
                                    dic_data_max[i]=max_length
                        min_length=df[i].dropna().astype(str).str.len().min()
                                    
                        if i not in  dic_data_min:
                            if str(min_length)=='nan':
                                dic_data_min[i]=0
                            else:
                                dic_data_min[i]=min_length
                        else:
                            if str(min_length)!='nan':
                                if dic_data_min[i]>min_length:
                                    dic_data_min[i]=min_length
                        if i not in unique_status:
                            unique_status[i]=True
                            unique_count[i]=set()
                        if not df[i].dropna().is_unique:
                            unique_status[i]=False
                        unique_count[i].update(df[i].dropna().unique())
                        if i not in not_null_count:
                            not_null_count[i]=df[i].count()
                        else:
                            not_null_count[i]+=df[i].count()
                        if i not in alpha_numeric:
                            if not df[i].isna().all():
                                filtered=df[df[i].str.contains(r'[a-zA-Z]+\s*[0-9]+\s*|\s*[0-9]+\s*[a-zA-Z]+',na=False)]
                                count_alpha=filtered.shape[0]
                                            
                                alpha_numeric[i]=count_alpha
                            else:
                                alpha_numeric[i]=0
                        else:
                            if not df[i].isna().all():
                                filtered=df[df[i].str.contains(r'[a-zA-Z]+\s*[0-9]+\s*|\s*[0-9]+\s*[a-zA-Z]+',na=False)]
                                count_alpha=filtered.shape[0]
                        
                                alpha_numeric[i]+=count_alpha
                            else:
                                alpha_numeric[i]+=0
                        if i not in numeric_count:
                            if not df[i].isna().all():
                                filtered_sum=df[df[i].str.match(r'^\d+$',na=False)]
                                count_num=filtered_sum.shape[0]
                                numeric_count[i]=count_num
                            else:
                                numeric_count[i]=0
                        else:
                            if not df[i].isna().all():
                                filtered_sum=df[df[i].str.match(r'^\d+$',na=False)]
                                count_num=filtered_sum.shape[0]
                                numeric_count[i]+=count_num
                            else:
                                numeric_count[i]+=0
                    unique_count={columns:len(value_set) for columns,value_set in unique_count.items()}
                    merg={key:[tot_rec,unique_count[key],dic_data_max[key],dic_data_min[key],not_null_count[key],tot_rec-not_null_count[key],round(100*float(not_null_count[key])/float(tot_rec),2),unique_status[key],alpha_numeric[key],'YES' if alpha_numeric[key]>0 else 'NO',numeric_count[key] ]for key in dic_data_max if key in dic_data_min and key in unique_status and key in unique_count}
                    main_df.update(merg)
                else:
                    width=df_config['Fixed_Width'].dropna().astype(int).tolist()
                    ch=[]
                    siz=200000
                    for e in encoding:
                        try:
                            cnt=1
                            for df in pd.read_fwf(path_1+'//'+file_name,widths=width,header=None,encoding=e,dtype=str,chunksize=siz):
                                df.columns=df_config['Columns'].dropna().tolist()
                                print('yes')
                                print(df.shape[0])
                                tot_rec=tot_rec+df.shape[0]
                                for i in df.columns: 
                                    df[i]=df[i].replace('nan',np.nan)
                                    max_length=df[i].dropna().astype(str).str.len().max()
                                    if i not in  dic_data_max:
                                        if str(max_length)=='nan':
                                            dic_data_max[i]=0
                                        else:
                                            dic_data_max[i]=max_length
                                    else:
                                        if str(max_length)!='nan':
                                            if dic_data_max[i]<max_length:
                                                dic_data_max[i]=max_length
                                    min_length=df[i].dropna().astype(str).str.len().min()
                                    
                                    if i not in  dic_data_min:
                                        if str(min_length)=='nan':
                                            dic_data_min[i]=0
                                        else:
                                            dic_data_min[i]=min_length
                                    else:
                                        if str(min_length)!='nan':
                                            if dic_data_min[i]>min_length:
                                                dic_data_min[i]=min_length
                                    if i not in unique_status:
                                        unique_status[i]=True
                                        unique_count[i]=set()
                                    if not df[i].dropna().is_unique:
                                        unique_status[i]=False
                                    unique_count[i].update(df[i].dropna().unique())
                                    if i not in not_null_count:
                                        not_null_count[i]=df[i].count()
                                    else:
                                        not_null_count[i]+=df[i].count()
                                    
                                    if i not in alpha_numeric:
                                        if not df[i].isna().all():
                                            filtered=df[df[i].str.contains(r'[a-zA-Z]+\s*[0-9]+\s*|\s*[0-9]+\s*[a-zA-Z]+',na=False)]
                                            count_alpha=filtered.shape[0]
                                            
                                            alpha_numeric[i]=count_alpha
                                        else:
                                            alpha_numeric[i]=0
                                    else:
                                        if not df[i].isna().all():
                                            filtered=df[df[i].str.contains(r'[a-zA-Z]+\s*[0-9]+\s*|\s*[0-9]+\s*[a-zA-Z]+',na=False)]
                                            count_alpha=filtered.shape[0]
                        
                                            alpha_numeric[i]+=count_alpha
                                        else:
                                            alpha_numeric[i]+=0
                                    if i not in numeric_count:
                                        if not df[i].isna().all():
                                            filtered_sum=df[df[i].str.match(r'^\d+$',na=False)]
                                            count_num=filtered_sum.shape[0]
                                            numeric_count[i]=count_num
                                        else:
                                            numeric_count[i]=0
                                    else:
                                        if not df[i].isna().all():
                                            filtered_sum=df[df[i].str.match(r'^\d+$',na=False)]
                                            count_num=filtered_sum.shape[0]
                                            numeric_count[i]+=count_num
                                        else:
                                            numeric_count[i]+=0
                            unique_count={columns:len(value_set) for columns,value_set in unique_count.items()}      
                            merg={key:[tot_rec,unique_count[key],dic_data_max[key],dic_data_min[key],not_null_count[key],tot_rec-not_null_count[key],round(100*float(not_null_count[key])/float(tot_rec),2),unique_status[key],alpha_numeric[key],'YES' if alpha_numeric[key]>0 else 'NO',numeric_count[key] ]for key in dic_data_max if key in dic_data_min and key in unique_status and key in unique_count}
                            main_df.update(merg)
                            break
                        except UnicodeDecodeError:
                            print('failed for ')
                #df1=df['SUBSCRIBER_SSN'].dropna()
                #df1.to_excel(path+'//'+file_nm+'.xlsx')
                graph=[("Atributes","Data_Percentage")]
                for key in dic_data_max:
                     graph_couple=(key,round(100*float(not_null_count[key])/float(tot_rec),2))
                     graph.append(graph_couple)
                main_df1=pd.DataFrame(main_df)
                #print(graph)
                wb=Workbook()
                ws=wb.active
                ws.title='Results'
                ws2=wb.create_sheet(title='Graphical Representation')
                ws.sheet_view.showGridLines=False
                ws2.sheet_view.showGridLines=False
                start_row=16
                wb.active=1
                for rw_idx,rw_data in enumerate(graph,start=start_row):
                    for col_idx,value in enumerate(rw_data,start=1):
                        ws2.cell(row=rw_idx,column=col_idx,value=value)
                chart=BarChart()
                chart.title="Visual Representation"
                data_range=Reference(ws2,min_col=2,min_row=start_row+1,max_col=2,max_row=start_row+len(graph)-1)
                category=Reference(ws2,min_col=1,min_row=start_row+1,max_row=start_row+len(graph)-1)
                chart.add_data(data_range,titles_from_data=False)
                
                chart.x_axis.majorGridlines=None
                chart.y_axis.majorGridlines=None
                chart_width=4+0.9*(len(graph))
                chart.width=chart_width
                chart.set_categories(category)
                #print(chart)
                chart.dataLabels=openpyxl.chart.label.DataLabelList()
                chart.dataLabels.showVal=True
                chart.x_axis.title="Attributes"
                #chart.y_axis.title.tx.rich.="Data Percentage"
                chart.y_axis.delete= True
                ws2.add_chart(chart,"A1")
                cell_1=ws2.cell(row=16,column=1)
                cell_2=ws2.cell(row=16,column=2)
                cell_1.font=Font(bold=True)
                cell_2.font=Font(bold=True)
                cell_1.fill=PatternFill(start_color="ADD8E6",end_color="ADD8E6",fill_type="solid")
                cell_2.fill=PatternFill(start_color="ADD8E6",end_color="ADD8E6",fill_type="solid")
                wb.active=0
                for c_id,head in enumerate(main_df1.columns,start=1):
                    cell=ws.cell(row=5,column=c_id+1,value=head)
                    cell.font=Font(bold=True)
                    cell.border=Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'))
                for r_id,row in enumerate(main_df1.values,start=6):
                    for c_id,value in enumerate(row,start=1):
                        cell=ws.cell(row=r_id,column=c_id+1,value=value)
                        if r_id==15:
                            if value=='YES':
                                cell.fill=PatternFill(start_color="D8D42C",end_color="D8D42C",fill_type="solid")
                        if r_id==12 and c_id+1>=3:
                            if int(value)<50:
                                cell.fill=PatternFill(start_color="FF7F7F",end_color="FF7F7F",fill_type="solid")
                        cell.border=Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'))
                for cell in ws['B'][2:]:
                    cell.font=Font(bold=True)
                
                cell_tot=ws.cell(row=2,column=1,value="Filename | Count ")
                cell_tot.font=Font(size=14,bold=True,color="00B0F0")
                #cell_tot.fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
                
                
                cell_scenario=ws.cell(row=5,column=2)
                cell_scenario.fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
                cell_val=ws.cell(row=3,column=1,value=file_name.split('.')[0] +' | '+str(tot_rec))
                #cell_val.font=Font(size=14,bold=True)
                wb.save(path+'//'+'DQ_Check_'+file_name.split('.')[0]+'.xlsx')
                wb.close()
                wb1 = xw.Book(path+'//'+'DQ_Check_'+file_name.split('.')[0]+'.xlsx')
                wb1.app.visible=False
                # Autofit columns
                for sheet in wb1.sheets:
                    sheet.autofit("columns")
                wb1.sheets[0].range('A:A').column_width=8
                wb1.save(path+'//'+'DQ_Check_'+file_name.split('.')[0]+'.xlsx')
                wb1.close()
                print('one done '+file_name)
                #tme.sleep(2)
                file_log.append(file_name)
                file_error.append('No error : Success')
        except FileNotFoundError:
            file_log.append(file_name)
            file_error.append('FileNotFoundError : Config file not found')
        except KeyError as e:
            file_log.append(file_name)
            file_error.append(str(e) +' Column name is incorrect or not found in config file')
            
    log={'File_Name':file_log,'Error':file_error}
    df_log=pd.DataFrame(log)
    
    df_log.to_excel(Log_Path+'//'+cur_date_time_f+'.xlsx')
    wb2=xw.Book(Log_Path+'//'+cur_date_time_f+'.xlsx')
    wb2.app.visible=False
    
    for sheet in wb2.sheets:
        #sheet.api.DisplayGridlines=False
        sheet.autofit("columns")
    wb2.save(Log_Path+'//'+cur_date_time_f+'.xlsx')
    wb2.close()
    filt_file=[file_nm for file_nm in os.listdir(path) if file_nm.startswith('DQ_Check') and file_nm.endswith('.xlsx')]
    for f in filt_file:
        src=os.path.join(path,f)
        dest=os.path.join(New_Val_Res_Path,f)
        shutil.move(src,dest)